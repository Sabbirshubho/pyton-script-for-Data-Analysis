# -*- coding: utf-8 -*-
"""
Created on Thu Jun 18 04:55:47 2020

@author: Sabbir
"""
import selenium
from selenium import webdriver
import pandas as pd  #panda 0.23.4
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
chrome_options = Options()
op=webdriver.ChromeOptions()
op.add_argument('headless')
chrome_options.add_argument('--no-sandbox')
writer = pd.ExcelWriter(r'Data.xlsx', engine='xlsxwriter')
writer.save()
#url='https://www.w3schools.com/html/html_tables.asp'
numeric_value=[6679,6685]
#numeric_value=[6685,6812]
web_count=2


for r in range(web_count):
    try:
        # https://intranet-grid.hms.se/intra/scala/pl/pl_pending.php?PL01001=6679&scco=HS
        url='-----------' #first part of url
        url+=str(numeric_value[r])
        url+='----'
        driver = selenium.webdriver.Chrome(options=op)
        driver.get(url)
        driver.implicitly_wait(30)
        all_tables=pd.read_html(driver.page_source, attrs={'class': 'sellist'})
        df = all_tables[0]
        print(df)
        row=len(df)
        driver.close()
        driver.quit()

    except Exception as e:
        print(e)
    print(row)
    #get all href in a webpage
    """
    elems = driver.find_elements_by_xpath("//a[@href]")
    for elem in elems:
        print(elem.get_attribute("href"))
    """
    all_Tables=[]
    Df=[] # sub page dataframe
    for x in range(row-1):
        print(x)
        try:
            all_Tables.append(x)
            Df.append(x)
            #//*[@id="row_0"]/td[1]/a
            #//*[@id="row_1"]/td[1]/a
            xpth='//*[@id="row_'
            #xpth='row_'
            xpth+=str(x)
            xpth+='"]/td[1]/a'
            #print(xpth)
            driver = selenium.webdriver.Chrome(options=op)
            
            driver.get(url)
            driver.maximize_window()
            driver.implicitly_wait(20)
            #driver.execute_script("arguments[0].click();", WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH,"xpth"))).element.click()
            
            #driver.find_element_by_id(xpth).click()
            # if x==x:
            #     sleep(10)
            
            element = driver.find_element_by_xpath(xpth)
            driver.execute_script("arguments[0].click();", element)
            #driver.find_element_by_xpath(xpth).click()
            all_Tables[x]=pd.read_html(driver.page_source, attrs={'class': 'sellist'})
            Df[x] = all_Tables[x]
            #print(Df[x])
            driver.close()
            driver.quit()

        except Exception as e:
            print(e)
    
   
    k=1
    m=0
    for x in range(row-1):
        try:
            df1=Df[x]
            y=df1[0]
            skp=len(y)
            rp=0
            m=m+1
            for z in range(1,skp+1):
                line = pd.DataFrame({}, index=[z])
                df = pd.concat([df.iloc[:z+m], line, df.iloc[z+m:]]).reset_index(drop=True)
            m=m+skp
        except Exception as e:
            print(e)
    # savefile = asksaveasfilename(filetypes=(("Excel files", "*.xlsx"),
    #                                     ("All files", "*.*") ))    
    #path = r'C:/Users/sash/Desktop/Importtable/Data.xlsx' # give path for ecell
    book = load_workbook('Data.xlsx')
    writer = pd.ExcelWriter('Data.xlsx', engine = 'openpyxl')
    writer.book = book
    
    rp=0
    for x in range(row-1):
        try:
            df1=Df[x]
            y=df1[0]
            if x==0:
                rp=1
            else:
                rp=rp+skp+1   
            # Position the dataframes in the worksheet.
            y.to_excel(writer, sheet_name=str(numeric_value[r]), startrow=rp, startcol=7,header=False,index=False)  # Default position, cell A1.    
            skp=len(y)
        except Exception as e:
            print(e)
    
    df.to_excel(writer, sheet_name=str(numeric_value[r]),header=False,index=False)
    writer.save()
    writer.close()
    
#path = r'C:/Users/sash/Desktop/Importtable/Data.xlsx'   # give path for exell
book = load_workbook('Data.xlsx')
book.get_sheet_names()
std=book.get_sheet_by_name('Sheet1')
writer = pd.ExcelWriter('Data.xlsx', engine = 'openpyxl')
writer.book = book    
book.remove_sheet(std)

book.save('Data.xlsx')

# wb.save('Data.xlsx')
